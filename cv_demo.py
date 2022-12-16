# -*- coding: utf-8 -*-
"""
@Time ： 2022/11/28 16:11
@Auth ： 蔍鸣
@File ：cv_demo.py
@IDE ：PyCharm
@Motto: 拔丝kite博
"""


from openpyxl import load_workbook
import cv2 as cv
from openpyxl.styles import PatternFill

#
image = cv.imread("./data/ikun.png", 1)
B = image[:, :, 0]
G = image[:, :, 1]
R = image[:, :, 2]
rowsCont = len(B[0])
colmCout = len(B)
#print(rowsCont, colmCout)

fileName = r"C:\Users\qchd\PycharmProjects\pythonProject\data\drawExcel.xlsx"
wb = load_workbook(fileName)
sheet = wb.worksheets[0]

for c in range(1, len(B[0])):
    # 遍历每一列的像素值
    for r in range(1, len(B)):
        # 遍历每一行的像素值
        # color = cel.fill.fgColor.rgb

        d = str(hex(R[r][c])).lstrip("0x").upper()
        e = str(hex(G[r][c])).lstrip("0x").upper()
        f = str(hex(B[r][c])).lstrip("0x").upper()
        while len(d) < 2: d = str(0) + d
        while len(e) < 2: e = str(0) + e
        while len(f) < 2: f = str(0) + f
        RGB = d + e + f

        fill = PatternFill(fill_type='solid', fgColor=RGB, bgColor="FFFFFF")
        sheet.cell(row=r, column=c).fill = fill
"""
"""
wb.save(fileName)
#
# #
print("\n==== I am great !! ========")




